from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_tabular_file(file_path: str | Path, file_type: str) -> tuple[list[Any], list[list[Any]]]:
    path = Path(file_path)

    if file_type == "csv":
        return _read_csv(path)
    if file_type == "xlsx":
        return _read_xlsx(path)

    raise ValueError("unsupported file type")


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def trim_blank_rows(rows: list[list[Any]]) -> list[list[Any]]:
    return [list(row) for row in rows if any(not is_blank(value) for value in row)]


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def is_numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return False
        try:
            float(stripped)
        except ValueError:
            return False
        return True
    return False


def to_float(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        return float(value.strip())
    raise ValueError("value is not numeric")


def _read_csv(file_path: Path) -> tuple[list[Any], list[list[Any]]]:
    raw_bytes = file_path.read_bytes()
    decoded_text: str | None = None

    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded_text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if decoded_text is None:
        raise ValueError("csv decoding failed")

    sample = decoded_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    rows = list(csv.reader(decoded_text.splitlines(), dialect=dialect))
    if not rows:
        return [], []

    header = rows[0]
    body = [list(row) for row in rows[1:]]
    return list(header), body


def _read_xlsx(file_path: Path) -> tuple[list[Any], list[list[Any]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], []

    header = list(rows[0])
    body = [list(row) for row in rows[1:]]
    return header, body
